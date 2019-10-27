"""
======================
@author:LiXuejin

@time:2019/10/26:21:20

@email:lixuejin@fang.com
======================
"""
from openpyxl import load_workbook

class TestCases:
    pass

class ReadExcel:

    def __init__(self,filename, sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name

    def open(self):
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def read_excel(self):
        self.open()
        cases = []
        rows = list(self.ws.rows)

        title = []
        for i in rows[0]:
            title.append(i.value)

        for r in rows[1:]:
            items = []
            for c in r:
                items.append(c.value)
            case_data = dict(zip(title, items))
            cases.append(case_data)


        datas = []
        for case in cases:
            case_obj = TestCases()
            for k, v in case.items():
                setattr(case_obj, k, v)
            datas.append(case_obj)
        return datas

    def write(self, row, column, value):
        self.ws.cell(row=row, column=column, value=value)
        self.wb.save(self.filename)
        pass


if __name__ == '__main__':
    do_excel = ReadExcel(r"D:\Program Files\python_23\unittest_exercise\data\register_data.xlsx","sheet")
    data = do_excel.read_excel()
    print(data[1].__dict__)
